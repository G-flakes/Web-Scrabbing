from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import pandas as pd
import time
import re
from datetime import datetime

# Check if the page has fully loaded
def check_page_loaded(driver):
    return driver.execute_script("return document.readyState") == "complete"

# Retry clicking a button up to 3 times if it fails
def retry_click(driver, locator):
    attempts = 0
    while attempts < 3:
        try:
            btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(locator))
            driver.execute_script("arguments[0].click();", btn)
            print("Click successful.")
            break
        except Exception as e:
            print(f"Retry clicking, attempt {attempts + 1}: {e}")
            time.sleep(1)
            attempts += 1

# Shorten text to 255 characters if longer
def paraphrase(text):
    return text if len(text) <= 255 else text[:255] + '...'

# Calculate the end-of-life date from the launch date and lifetime
def calculate_end_of_life(launch_date, lifetime):
    if "cancelled" in launch_date.lower():
        return "cancelled"
    try:
        years_to_add = int(re.search(r'\d+', lifetime).group())
        if '.' in launch_date:
            launch_date_dt = datetime.strptime(launch_date, "%d.%m.%Y")
            end_of_life_dt = launch_date_dt.replace(year=launch_date_dt.year + years_to_add)
            return end_of_life_dt.strftime("%d.%m.%Y")
        else:
            return str(int(launch_date) + years_to_add)
    except Exception as e:
        return "Lifetime data not found"

# Fetch the orbit type from the webpage
def get_orbit_type(driver):
    try:
        orbit_type_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="sdorb"]'))
        )
        return orbit_type_element.text
    except Exception as e:
        return "Orbit type not available"

# Fetch the mass of the satellite
def get_mass(driver):
    try:
        mass_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="sdmas"]'))
        )
        return float(re.search(r'\d+\.?\d*', mass_element.text).group())
    except Exception as e:
        return "Mass data not available"

# Determine the mission status based on launch and end-of-life dates and background info
def determine_msd(launch_date, end_of_life, background_info):
    if "cancelled" in launch_date.lower() or "cancelled" in end_of_life.lower():
        return "Cancelled"
    elif "mission complete" in background_info.lower() or "failure" in background_info.lower():
        return "Mission complete"
    try:
        launch_date_dt = datetime.strptime(launch_date, "%d.%m.%Y")
        if datetime.now() < launch_date_dt:
            return "Planned"
    except ValueError:
        if int(launch_date) > datetime.now().year:
            return "Planned"
    try:
        end_of_life_dt = datetime.strptime(end_of_life, "%d.%m.%Y")
        if datetime.now() < end_of_life_dt:
            return "Operational"
    except:
        pass
    if "extended" in background_info.lower():
        return "Extended Mission"
    return "Operational"  # Default status

# Determine potential product candidates based on mass, orbit type, and additional criteria
def determine_ppc(mass, orbit_type, end_of_life, background_info):
    ppc = []
    
    # Define keywords for MEV and MRV
    mev_keywords = ['life extension', 'altitude control', 'adjustment', 'orbital adjustment', 
                    'propulsion support', 'payload', 'payload delivery', 'relocation']
    mrv_keywords = ['repair', 'debris', 'servicing', 'maintenance', 'refueling', 'payload', 'payload delivery']
    
    try:
        # Calculate remaining years to determine if MEV is needed
        remaining_years = (end_of_life - datetime.now()).days / 365.25
        if any(keyword in background_info.lower() for keyword in mev_keywords) and "GEO" in orbit_type:
            ppc.append("MEV")
    except Exception as e:
        print("Error processing end_of_life date:", e)  # Print error message to understand what went wrong

    # Check if MRV is required based on background info keywords
    if any(keyword in background_info.lower() for keyword in mrv_keywords):
        ppc.append("MRV")

    # Determine if MEP is suitable based on mass and orbit type
    if 1500 < mass < 2500 and "GEO" in orbit_type:
        ppc.append("MEP")
    return ', '.join(ppc) if ppc else "No P.P.C determined"

# Setup the WebDriver and open the target URL
service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.get("https://space.skyrocket.de/index.html")

# Accept cookies on the page
retry_click(driver, (By.CSS_SELECTOR, ".cc_btn.cc_btn_accept_all"))

# Search for 'geostationary' satellites
try:
    search_box = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "searchbox")))
    search_box.clear()
    search_box.send_keys("geostationary" + Keys.ENTER)
    print("Search for 'geostationary' submitted.")
except Exception as e:
    print(f"Error during search: {e}")
    driver.quit()

# Initialize DataFrame to store the data
df = pd.DataFrame(columns=[
    "Link", "Satellite name", "Mission Type", "Background", "Photo description",
    "M.S.D (Mission Status Details)", "Launch Date", "End of life", "Orbit type",
    "Class type", "P.P.C (Potential Product Candidate)"
])

# Iterate through all satellite links starting from a specified index
links = driver.find_elements(By.XPATH, "/html/body/div[1]/div[1]/div/table/tbody/tr[2]/td/ul/li/a")
start_index = 29
for i in range(start_index, len(links)):
    try:
        href = links[i].get_attribute('href')
        driver.get(href)
        WebDriverWait(driver, 10).until(check_page_loaded)

        satellite_name = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="satlist"]/tbody/tr[2]/td[1]'))
        ).text
        mission_type = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="sdtyp"]'))
        ).text
        launch_date = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="satlist"]/tbody/tr[2]/td[3]'))
        ).text
        lifetime = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="sdlif"]'))
        ).text
        end_of_life = calculate_end_of_life(launch_date, lifetime)
        mass = get_mass(driver)
        orbit_type = get_orbit_type(driver)
        class_type = "Class A" if mass >= 1000 else "Class B" if mass >= 500 else "Class N"
        background_info = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "div#satdescription"))
        ).text
        background_summary = paraphrase(background_info)
        msd = determine_msd(launch_date, end_of_life, background_info)
        ppc = determine_ppc(mass, orbit_type, 10, background_info)  # Assume lifespan_remaining for example

        try:
            image_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, "//*[@id='contimg']/img"))
            )
            image_url = image_element.get_attribute('src')
        except:
            image_url = "No photo for this satellite is currently available"

        new_row = {
            "Link": href,
            "Satellite name": satellite_name,
            "Mission Type": mission_type,
            "Background": background_summary,
            "Photo description": image_url,
            "M.S.D (Mission Status Details)": msd,
            "Launch Date": launch_date,
            "End of life": end_of_life,
            "Orbit type": orbit_type,
            "Class type": class_type,
            "P.P.C (Potential Product Candidate)": ppc
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        print(f"Processed link {i - start_index + 1}: {href}, Satellite Name: {satellite_name}, "
              f"Mission Type: {mission_type}, Background: {background_summary}, "
              f"Image URL: {image_url}, Launch Date: {launch_date}, End of Life: {end_of_life}, "
              f"Orbit Type: {orbit_type}, Class Type: {class_type}, "
              f"M.S.D: {msd}, P.P.C: {ppc}")
        driver.back()
        WebDriverWait(driver, 10).until(check_page_loaded)
    except Exception as e:
        print(f"Failed processing link {i - start_index + 1}: {e}")
        driver.back()

# Save the data to Excel
try:
    # Save initial data
    df.to_excel("Gunthers_results.xlsx", index=False)
    print("Data collection complete and saved to 'Gunthers_results.xlsx'.")

    # Filter and create a revised sheet
    filtered_df = df[df['M.S.D (Mission Status Details)'] != 'Cancelled']
    filtered_df = filtered_df[~filtered_df['Launch Date'].str.contains('cancelled', na=False)]
    filtered_df = filtered_df[~filtered_df['End of life'].str.contains('cancelled', na=False)]
    filtered_df = filtered_df[filtered_df['Orbit type'].str.contains('GEO')]
    filtered_df = filtered_df[filtered_df['Class type'] == 'Class A']

    with pd.ExcelWriter("Gunthers_results.xlsx", engine='openpyxl', mode='a') as writer:
        filtered_df.to_excel(writer, sheet_name='Sheet1_Revised', index=False)
    print("Revised data also saved in 'Gunthers_results.xlsx'.")

except Exception as e:
    print("Failed to save data to Excel:", e)

driver.quit()

# Load the existing Excel workbook
wb = load_workbook('Gunthers_results.xlsx')

# Select a specific sheet to format
ws = wb['Sheet1_Revised']  # 'Sheet1_Revised' on the excel sheet goning to be formatted

# Define styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue fill
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                      top=Side(style='thick'), bottom=Side(style='thick'))

# Apply styles to the header row
for cell in ws['1:1']:  # Assuming the first row is the header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thick_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Apply borders to all other cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# Save the changes to the workbook
wb.save('Gunthers_results.xlsx')
